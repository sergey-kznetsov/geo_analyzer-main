from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from geo_analyzer.reporting.excel_report import export_report_to_excel


def test_excel_report_contains_new_accessibility_columns(tmp_path: Path):
    output_path = tmp_path / "report.xlsx"

    analysis_result = {
        "meta": {
            "resolved_address": "Ижевск, Пушкинская 277",
            "latitude": 56.8526,
            "longitude": 53.2115,
            "isochrones_minutes": [5, 10, 15],
            "provider": "2GIS",
        },
        "text_summary": "Тестовое саммари",
        "quality_scores": pd.DataFrame(
            [
                {
                    "Метрика": "Транспортная доступность",
                    "Оценка_из_10": 7.5,
                    "Пояснение": "Тест",
                    "Шкала_оценки": "0-10",
                }
            ]
        ),
        "accessibility_snapshot": pd.DataFrame(
            [
                {
                    "Зона_доступности": "0–5 мин",
                    "Минут_пешком": 5,
                    "Количество_POI": 10,
                    "Количество_категорий": 4,
                    "Остановочных_комплексов": 2,
                    "Точек_притяжения_городского_масштаба": 1,
                    "Пешая_доступность_из_10": 6.5,
                    "Остановочная_доступность_из_10": 5.0,
                    "Авто_доступность_до_центра_из_10": 8.5,
                    "Итоговая_доступность_из_10": 6.7,
                    "Авто_время_до_центра_мин": 8,
                    "Авто_расстояние_до_центра_км": 3.2,
                    "Источник_авто_метрики": "2gis_routing_api",
                }
            ]
        ),
        "category_summary": pd.DataFrame(
            [
                {
                    "Категория_2GIS": "Супермаркет",
                    "functional_category": "Повседневная торговля и услуги",
                    "Количество": 2,
                    "Доля_проц": 20,
                }
            ]
        ),
        "poi_details_by_iso": pd.DataFrame(
            [
                {
                    "Название": "Магнит",
                    "Адрес": "Пушкина, 1",
                    "Категория_2GIS": "Супермаркет",
                    "functional_category": "Повседневная торговля и услуги",
                    "Минут_пешком": 5,
                    "Зона_доступности": "0–5 мин",
                    "rubrics_2gis": ["Супермаркеты"],
                }
            ]
        ),
        "attraction_points": pd.DataFrame(),
        "anti_driver_summary": pd.DataFrame(),
        "benchmark_summary": pd.DataFrame(),
        "network_metrics": pd.DataFrame(),
        "city_benchmark": {
            "meta": {
                "city": "Ижевск",
                "data_source": "config_snapshot",
                "benchmark_version": "1",
            },
            "thresholds_city": {"Инфраструктурная насыщенность": 7},
            "thresholds_district": {},
            "weights": {},
        },
        "drive_metrics": {
            "drive_time_min": 8,
            "drive_distance_km": 3.2,
            "center_name": "Центральная площадь",
            "center_city": "Ижевск",
            "data_source": "2gis_routing_api",
        },
    }

    export_report_to_excel(
        analysis_result=analysis_result,
        visuals={},
        gamma_prompt="",
        output_path=output_path,
    )

    assert output_path.exists()

    workbook = load_workbook(output_path)
    expected_sheets = {
        "Саммари",
        "Качество среды",
        "POI по изохронам",
        "Точки притяжения",
        "Антидрайверы",
        "Бенчмарки",
        "Сетевые метрики",
    }
    assert expected_sheets.issubset(set(workbook.sheetnames))
    # Доступность теперь сворачивается в отдельную строку листа "Саммари",
    # а авто-метрики берутся из drive_metrics.
    summary_sheet = workbook["Саммари"]
    summary_headers = [cell.value for cell in summary_sheet[1]]
    section_idx = summary_headers.index("Раздел")
    sections = {row[section_idx].value for row in summary_sheet.iter_rows(min_row=2)}
    assert "Доступность" in sections
    assert "Авто-доступность" in sections