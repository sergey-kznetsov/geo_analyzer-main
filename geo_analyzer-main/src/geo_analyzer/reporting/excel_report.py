from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from geo_analyzer.core.utils import normalize_quality_scores, safe_value, to_numeric


TECHNICAL_COLUMNS = {
    "Рейтинг",
    "Количество_отзывов",
    "Источник",
    "dgis_id",
    "fid",
    "rubric_id",
    "rubrics_2gis",
    "source_categories_2gis",
    "source_category_2gis",
    "category_groups_2gis",
    "classification_rule_id",
    "classification_status",
    "validation_status",
    "raw_2gis",
    "geometry",
}


def _drop_technical_columns(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if df.empty:
        return df.copy()
    result = df.copy()
    drop_cols = [column for column in result.columns if str(column) in TECHNICAL_COLUMNS]
    return result.drop(columns=drop_cols, errors="ignore")


def _excel_safe_df(df: pd.DataFrame | None) -> pd.DataFrame:
    result = _drop_technical_columns(df)
    if result.empty:
        return result.copy()
    for column in result.columns:
        result[column] = result[column].map(safe_value)
    return result


def _drive_value(drive_metrics: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = drive_metrics.get(key)
        if value is not None:
            return value
    return pd.NA


def _row_value(row: pd.Series, *keys: str, default: Any = pd.NA) -> Any:
    for key in keys:
        value = row.get(key)
        try:
            if pd.isna(value):
                continue
        except (TypeError, ValueError):
            pass
        if value is not None:
            return value
    return default


def _prepare_summary_df(analysis_result: dict[str, Any], drive_metrics: dict[str, Any]) -> pd.DataFrame:
    meta = analysis_result.get("meta", {}) or {}
    quality_scores = normalize_quality_scores(analysis_result.get("quality_scores", pd.DataFrame()))
    category_summary = analysis_result.get("category_summary", pd.DataFrame())
    accessibility = analysis_result.get("accessibility_snapshot", pd.DataFrame())
    attraction_summary = analysis_result.get("attraction_summary", pd.DataFrame())
    anti_driver_summary = analysis_result.get("anti_driver_summary", pd.DataFrame())
    rows: list[dict[str, Any]] = [
        {"Раздел": "Общий вывод", "Показатель": "Текстовое саммари", "Значение": analysis_result.get("text_summary"), "Комментарий": "Краткая интерпретация результата анализа."},
        {"Раздел": "Вводные", "Показатель": "Адрес / метка", "Значение": meta.get("resolved_address"), "Комментарий": "Точка, по которой выполнен анализ."},
        {"Раздел": "Вводные", "Показатель": "Координаты", "Значение": f"{meta.get('latitude')}, {meta.get('longitude')}", "Комментарий": "Широта и долгота точки анализа."},
        {"Раздел": "Вводные", "Показатель": "Город", "Значение": meta.get("city"), "Комментарий": "Город, к которому привязан benchmark."},
        {"Раздел": "Вводные", "Показатель": "Изохроны", "Значение": ", ".join(map(str, meta.get("isochrones_minutes", []))), "Комментарий": "Непересекающиеся зоны доступности."},
        {"Раздел": "Вводные", "Показатель": "Провайдер данных", "Значение": meta.get("provider", "2GIS"), "Комментарий": "Источник POI и транспортных данных."},
        {"Раздел": "Вводные", "Показатель": "Шкала оценки", "Значение": "0-10", "Комментарий": "Чем выше значение, тем сильнее показатель."},
    ]

    if not quality_scores.empty:
        avg_score = round(float(quality_scores["Оценка_из_10"].mean()), 2)
        best_row = quality_scores.sort_values("Оценка_из_10", ascending=False).iloc[0]
        weak_row = quality_scores.sort_values("Оценка_из_10", ascending=True).iloc[0]
        rows.extend([
            {"Раздел": "Качество среды", "Показатель": "Средняя оценка", "Значение": avg_score, "Комментарий": "Среднее по всем метрикам качества среды."},
            {"Раздел": "Качество среды", "Показатель": "Сильнейшая метрика", "Значение": f"{best_row['Метрика']} — {best_row['Оценка_из_10']} из 10", "Комментарий": "Главная сильная сторона локации."},
            {"Раздел": "Качество среды", "Показатель": "Слабейшая метрика", "Значение": f"{weak_row['Метрика']} — {weak_row['Оценка_из_10']} из 10", "Комментарий": "Зона внимания для продуктовой интерпретации."},
        ])

    if category_summary is not None and not category_summary.empty and "Количество" in category_summary.columns:
        total_poi = int(to_numeric(category_summary["Количество"]).sum())
        rows.append({"Раздел": "Инфраструктура", "Показатель": "Всего POI", "Значение": total_poi, "Комментарий": "Количество объектов после загрузки, классификации и дедупликации."})

    if accessibility is not None and not accessibility.empty:
        acc = accessibility.copy()
        score_col = "Итоговая_доступность_из_10"
        if score_col in acc.columns:
            best_access = acc.sort_values(score_col, ascending=False).iloc[0]
            rows.append({"Раздел": "Доступность", "Показатель": "Лучшая зона", "Значение": f"{best_access.get('Зона_доступности', 'нет данных')} — {round(float(best_access.get(score_col, 0)), 2)} из 10", "Комментарий": "Лучшая зона по совокупной доступности."})

    if attraction_summary is not None and not attraction_summary.empty and {"Показатель", "Значение"}.issubset(attraction_summary.columns):
        row = attraction_summary[attraction_summary["Показатель"].astype(str).eq("Индекс притяжения, из 100")]
        if not row.empty:
            value = row.iloc[0].get("Значение")
            rows.append({"Раздел": "Притяжение", "Показатель": "Индекс притяжения", "Значение": f"{round(float(value) / 10, 2)} из 10" if pd.notna(value) else "нет данных", "Комментарий": "Сила локации как точки притяжения."})

    if anti_driver_summary is not None and not anti_driver_summary.empty:
        anti_text = "; ".join(
            f"{row['Тип_антидрайвера']} — {int(row['Количество'])}"
            for _, row in anti_driver_summary.head(5).iterrows()
            if "Тип_антидрайвера" in row.index and "Количество" in row.index
        )
        rows.append({"Раздел": "Антидрайверы", "Показатель": "Основные антидрайверы", "Значение": anti_text or "не выявлены", "Комментарий": "Факторы, снижающие качество среды."})
        if not total_row.empty:
            row = total_row.iloc[0]
            rows.extend([
                {"Раздел": "Парковочный потенциал", "Показатель": "Потенциал до 10 минут", "Значение": _row_value(row, "Парковочный_потенциал_из_10", "Оценка_из_10", "Парковочный_коэффициент"), "Комментарий": "Формула: взвешенные парковочные места / (квартиры × 0.8) × 10."},
                {"Раздел": "Парковочный потенциал", "Показатель": "Жилых домов до 10 минут", "Значение": _row_value(row, "Жилых_домов"), "Комментарий": "Количество физических type=building домов, попавших в изохрону до 10 минут."},
                {"Раздел": "Парковочный потенциал", "Показатель": "Проверенных карточек домов", "Значение": _row_value(row, "Домов_с_проверенной_карточкой_2GIS"), "Комментарий": "Сколько домов прошло дозапрос карточки 2GIS по id."},
                {"Раздел": "Парковочный потенциал", "Показатель": "Квартир до 10 минут", "Значение": _row_value(row, "Квартир_в_зоне"), "Комментарий": "Сумма точных и подтверждённо оценочных квартир по жилым домам."},
                {"Раздел": "Парковочный потенциал", "Показатель": "Парковочных мест до 10 минут", "Значение": _row_value(row, "Парковочных_мест"), "Комментарий": "Только включённые в расчёт парковочные места."},
                {"Раздел": "Парковочный потенциал", "Показатель": "Точных мест 2GIS", "Значение": _row_value(row, "Парковочных_мест_точных_2GIS"), "Комментарий": "Места из capacity/атрибутов 2GIS."},
                {"Раздел": "Парковочный потенциал", "Показатель": "Оценочных мест", "Значение": _row_value(row, "Парковочных_мест_оценочных"), "Комментарий": "Места, оценённые только после проверки карточки 2GIS."},
                {"Раздел": "Парковочный потенциал", "Показатель": "Класс", "Значение": _row_value(row, "Класс_парковочного_потенциала", "Класс_обеспеченности"), "Комментарий": "8-10 высокий, 4-7 средний, 0-3 низкий."},
            ])

    rows.extend([
        {"Раздел": "Авто-доступность", "Показатель": "Центр для оценки", "Значение": drive_metrics.get("center_name") or "нет данных", "Комментарий": "Центр, до которого считается автомобильная доступность."},
        {"Раздел": "Авто-доступность", "Показатель": "Город центра", "Значение": drive_metrics.get("center_city") or "нет данных", "Комментарий": "Город, для которого определён центр."},
        {"Раздел": "Авто-доступность", "Показатель": "Авто-время до центра, мин", "Значение": _drive_value(drive_metrics, "drive_time_min", "avg_drive_time_min", "time_min"), "Комментарий": "Время на автомобиле до центра города."},
        {"Раздел": "Авто-доступность", "Показатель": "Авто-расстояние до центра, км", "Значение": _drive_value(drive_metrics, "drive_distance_km", "avg_drive_distance_km", "distance_km"), "Комментарий": "Расстояние на автомобиле до центра города."},
        {"Раздел": "Авто-доступность", "Показатель": "Источник авто-метрики", "Значение": drive_metrics.get("data_source"), "Комментарий": "2GIS Routing API, кеш или fallback."},
    ])

    return pd.DataFrame(rows)


def _prepare_poi_iso_df(poi_details_by_iso: pd.DataFrame | None) -> pd.DataFrame:
    columns = ["Название", "Адрес", "Категория_2GIS", "functional_category", "Минут_пешком", "Зона_доступности"]
    if poi_details_by_iso is None or poi_details_by_iso.empty:
        return pd.DataFrame(columns=columns)
    data = poi_details_by_iso.copy()
    if "Категория_2GIS" not in data.columns:
        data["Категория_2GIS"] = data.get("Категория", "Прочее")
    if "functional_category" not in data.columns:
        data["functional_category"] = data.get("Сценарная_группа", "Прочее")
    for column in columns:
        if column not in data.columns:
            data[column] = pd.NA
    data["Минут_пешком"] = pd.to_numeric(data["Минут_пешком"], errors="coerce").astype("Int64")
    return data[columns].drop_duplicates().sort_values(["Минут_пешком", "Категория_2GIS", "Название"], ascending=[True, True, True]).reset_index(drop=True)


def _prepare_attraction_points_df(attraction_points: pd.DataFrame | None) -> pd.DataFrame:
    columns = ["Название", "Адрес", "Категория_2GIS", "Функциональная_группа", "Минут_пешком", "Зона_доступности", "Городской_масштаб", "Балл_притяжения_из_10", "Статус_объекта"]
    if attraction_points is None or attraction_points.empty:
        return pd.DataFrame(columns=columns)
    data = attraction_points.copy()
    if "Категория_2GIS" not in data.columns:
        data["Категория_2GIS"] = data.get("Категория", "Прочее")
    if "Функциональная_группа" not in data.columns:
        data["Функциональная_группа"] = data.get("functional_category", pd.NA)
    if "Балл_притяжения_из_10" not in data.columns and "Балл_притяжения_из_100" in data.columns:
        data["Балл_притяжения_из_10"] = pd.to_numeric(data["Балл_притяжения_из_100"], errors="coerce") / 10
    for column in columns:
        if column not in data.columns:
            data[column] = pd.NA
    return data[columns].sort_values(["Балл_притяжения_из_10", "Минут_пешком", "Название"], ascending=[False, True, True], na_position="last").reset_index(drop=True)


def _prepare_anti_driver_df(anti_driver_summary: pd.DataFrame | None) -> pd.DataFrame:
    columns = ["Тип_антидрайвера", "Группа_антидрайвера", "Количество", "Суммарный_штраф", "Средний_штраф", "Пояснение"]
    if anti_driver_summary is None or anti_driver_summary.empty:
        return pd.DataFrame(columns=columns)
    data = anti_driver_summary.copy()
    for column in columns:
        if column not in data.columns:
            data[column] = pd.NA
    return data[columns].sort_values(["Суммарный_штраф", "Количество"], ascending=[False, False]).reset_index(drop=True)
def _prepare_network_metrics_df(network_metrics: pd.DataFrame | None) -> pd.DataFrame:
    columns = ["Метрика", "Значение", "Пояснение", "Шкала_оценки"]
    if network_metrics is None or network_metrics.empty:
        return pd.DataFrame(columns=columns)
    data = network_metrics.copy()
    for column in columns:
        if column not in data.columns:
            data[column] = pd.NA
    data["Значение"] = pd.to_numeric(data["Значение"], errors="coerce").clip(0, 10).round(2)
    data["Шкала_оценки"] = "0-10"
    return data[columns]


def _prepare_benchmark_df(benchmark_summary: pd.DataFrame | None) -> pd.DataFrame:
    columns = ["Метрика", "Фактическая_оценка", "Бенчмарк_района", "Бенчмарк_города", "Бенчмарк_города_из_10", "Отклонение_от_города", "Источник_городского_бенча", "Пояснение"]
    if benchmark_summary is None or benchmark_summary.empty:
        return pd.DataFrame(columns=columns)
    data = benchmark_summary.copy()
    for column in columns:
        if column not in data.columns:
            data[column] = pd.NA
    return data[columns]


def _style_workbook(output_path: Path) -> None:
    wb = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap_alignment
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap_alignment
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
        ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def export_report_to_excel(
    analysis_result: dict[str, Any],
    visuals: dict[str, Any],
    gamma_prompt: str,
    output_path: Path,
) -> Path:
    del visuals
    del gamma_prompt
    output_path.parent.mkdir(parents=True, exist_ok=True)
    drive_metrics = analysis_result.get("drive_metrics", {}) or {}

    sheets = {
        "Саммари": _prepare_summary_df(analysis_result, drive_metrics),
        "Качество среды": normalize_quality_scores(analysis_result.get("quality_scores", pd.DataFrame())),
        "POI по изохронам": _prepare_poi_iso_df(analysis_result.get("poi_details_by_iso", pd.DataFrame())),
        "Точки притяжения": _prepare_attraction_points_df(analysis_result.get("attraction_points", pd.DataFrame())),
        "Антидрайверы": _prepare_anti_driver_df(analysis_result.get("anti_driver_summary", pd.DataFrame())),
        "Бенчмарки": _prepare_benchmark_df(analysis_result.get("benchmark_summary", pd.DataFrame())),
        "Сетевые метрики": _prepare_network_metrics_df(analysis_result.get("network_metrics", pd.DataFrame())),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            _excel_safe_df(df).to_excel(writer, sheet_name=sheet_name[:31], index=False)

    _style_workbook(output_path)
    return output_path
